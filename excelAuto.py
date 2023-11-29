import tkinter as tk
from tkinter import ttk
import openpyxl
import os

class ProgramaExcel:
    def __init__(self, root):
        # Configuración inicial de la interfaz gráfica
        self.root = root
        self.root.title("Programa Excel")

        # Crear un cuadro de entrada y un botón para agregar
        self.entry_var = tk.StringVar()
        self.entry = ttk.Entry(root, textvariable=self.entry_var, width=30)
        self.entry.grid(row=0, column=0, padx=10, pady=10)

        self.add_button = ttk.Button(root, text="Agregar a Excel", command=self.escribir_en_excel)
        self.add_button.grid(row=0, column=1, padx=10, pady=10)

        # Crear un cuadro de entrada y un botón para eliminar
        self.delete_var = tk.StringVar()
        self.delete_entry = ttk.Entry(root, textvariable=self.delete_var, width=30)
        self.delete_entry.grid(row=1, column=0, padx=10, pady=10)

        self.delete_button = ttk.Button(root, text="Eliminar de Excel", command=self.eliminar_de_excel)
        self.delete_button.grid(row=1, column=1, padx=10, pady=10)

        # Crear un Listbox para la vista previa
        self.preview_listbox = tk.Listbox(root, height=10, width=40)
        self.preview_listbox.grid(row=0, column=2, rowspan=2, padx=10, pady=10)

        # Ruta completa al archivo Excel
        self.archivo_excel = os.path.join(os.getcwd(), 'datos.xlsx')

        # Actualizar la vista previa al iniciar
        self.actualizar_vista_previa()

    def escribir_en_excel(self):
        # Obtener el texto ingresado en el cuadro de entrada
        texto = self.entry_var.get()

        # Cargar o crear el archivo Excel
        try:
            libro = openpyxl.load_workbook(self.archivo_excel)
            hoja = libro.active
        except FileNotFoundError:
            libro = openpyxl.Workbook()
            hoja = libro.active
            hoja.append(['Texto'])  # Agregar un encabezado si el archivo no existe

        # Agregar el texto a la hoja de Excel
        hoja.append([texto])

        # Guardar el archivo Excel
        libro.save(self.archivo_excel)

        # Limpiar la entrada después de agregar a Excel
        self.entry_var.set("")

        # Actualizar la vista previa
        self.actualizar_vista_previa()

    def eliminar_de_excel(self):
        # Obtener el texto ingresado para eliminar
        texto_a_eliminar = self.delete_var.get()

        # Cargar el archivo Excel
        try:
            libro = openpyxl.load_workbook(self.archivo_excel)
            hoja = libro.active
        except FileNotFoundError:
            # Si el archivo no existe, no hay nada que eliminar
            return

        # Buscar y eliminar el texto de la hoja de Excel
        for fila in hoja.iter_rows(min_row=2, max_col=1, max_row=hoja.max_row):
            if fila[0].value == texto_a_eliminar:
                hoja.delete_rows(fila[0].row)
                break

        # Guardar el archivo Excel
        libro.save(self.archivo_excel)

        # Limpiar la entrada después de eliminar de Excel
        self.delete_var.set("")

        # Actualizar la vista previa
        self.actualizar_vista_previa()

    def actualizar_vista_previa(self):
        # Cargar el archivo Excel y obtener los datos
        try:
            libro = openpyxl.load_workbook(self.archivo_excel)
            hoja = libro.active
            datos = [fila[0].value for fila in hoja.iter_rows(min_row=2, max_col=1, max_row=hoja.max_row)]
        except FileNotFoundError:
            # Si el archivo no existe, no hay datos para mostrar
            datos = []

        # Limpiar el Listbox y agregar los datos
        self.preview_listbox.delete(0, tk.END)
        for dato in datos:
            self.preview_listbox.insert(tk.END, dato)

if __name__ == "__main__":
    # Crear la ventana principal
    root = tk.Tk()

    # Inicializar la instancia de ProgramaExcel
    programa_excel = ProgramaExcel(root)

    # Iniciar el bucle principal de la interfaz gráfica
    root.mainloop()
